from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

data = load_workbook("testing.xlsx")

analysis = data["Analysis"]

normal_cols = [4, 7]
normalize_cols = [9, 10, 24, 25, 31, 32]
row_range = [2, 69]
starting_col = 34


for i in range(8):
  ws = data[data.sheetnames[i]]
  for col in range(len(normalize_cols)):
    column_letter = get_column_letter(normalize_cols[col])
    range_str = column_letter + str(row_range[0]) + ":" + column_letter + str(row_range[1])
    for row in range(row_range[0], row_range[1]+1):
      # val = ws.cell(row, normalize_cols[col]).value
      val = column_letter + str(row)
      formula = "=(" + val + "-MIN(" + range_str + "))/(MAX(" + range_str + ")-MIN(" + range_str + "))"
      print(formula)
      # =(D2-MIN(D2:D69))/(MAX(D2:D69)-MIN(D2:D69))
      ws.cell(row, starting_col + col).value = formula
 

data.save("testing.xlsx")