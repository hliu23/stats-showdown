# run with "py check_results.py"

# loading data and input
from openpyxl import load_workbook
import sys
data = load_workbook("NCAABBStats2013-2021BDAB2021AlgorithmDevelopment.xlsx", read_only=True)
if (len(sys.argv) > 1):
  ws = data[sys.argv[1]]
else:
  ws = data["2021"]

teams = []
with open("input.txt", "r") as f:
  for line in f: 
    teams.append(line.strip())


# calculating results
name_col = 1
cind_col = 4
wins_col = 6

total = 0

for i in range(10):
  team_rank = 10 - i
  for row_num in range(1, ws.max_row):
    if ws.cell(row=row_num, column=name_col).value.strip() == teams[i]:
      wins = ws.cell(row_num, wins_col).value
      cind = ws.cell(row_num, cind_col).value
      total += team_rank * wins
      if cind == 1:
        total += (5 + wins * 5) * wins / 2

print(total)