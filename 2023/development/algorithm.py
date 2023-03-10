# TO RUN: install python and openpyxl package
from openpyxl import load_workbook
from dotenv import load_dotenv
import os

load_dotenv()
WORKBOOK_NAME = os.environ.get("WORKBOOK_NAME")

# TO RUN: replace workbook name and worksheet name

WB = load_workbook(WORKBOOK_NAME, read_only=True)
WS = WB["2022"]

# TO RUN: execute with "py algorithm.py"

# load data from workbook into nested list
COLS_OF_INTEREST = [1, 4, 7, 9, 10, 23, 24, 25, 27, 31]
NORMALIZE_STARTING = 3
CORREL_FACTORS = [-7.369519723, 3.207527431, 2.098710386, -1.634276411, 1.215039434, -3.164313497, 4.014363826, 0.742103775, 2.474446861]

data = []
for row in range(2, 69):
  col_list = []
  for col in COLS_OF_INTEREST:
    col_list.append(WS.cell(row, col).value)
  data.append(col_list)

# find max and min of a column
min_max = []
for col in range(NORMALIZE_STARTING, len(data[0])):
  min = data[0][col]
  max = data[0][col]
  for row in range(1, len(data)):
    val = data[row][col]
    if val < min: 
      min = val
    elif val > max:
      max = val
  min_max.append((min, max))


# normalize data to number between 0 and 1 if applicable
for col in range(NORMALIZE_STARTING, len(data[0])):
  for row in range(len(data)):
    normalized = (data[row][col] - min_max[col-3][0]) / (min_max[col-3][1] - min_max[col-3][0])
    data[row][col] = normalized

# multiply data by correlation efficients calculated ahead of time and sum results for a row
sort_list = []
for row in range(len(data)):
  result = 0
  for col in range(1, len(data[row])):
    val = data[row][col]
    result += CORREL_FACTORS[col-1] * val
  sort_list.append((data[row][0], result))

# sort the resultant list
def comp(e):
  return e[1]

sort_list.sort(reverse=True, key=comp)
for i in range(10):
  print(sort_list[i][0])

