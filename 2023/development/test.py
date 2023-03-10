from openpyxl import load_workbook
from dotenv import load_dotenv
import os
import statistics as st

load_dotenv()
WORKBOOK_NAME = os.environ.get("WORKBOOK_NAME")
WB = load_workbook(WORKBOOK_NAME, read_only=True)

# COLS_OF_INTEREST = [1, 4, 7, 9, 10, 23, 24, 25, 31, 32]
# NORMALIZE_STARTING = 2
# CORREL_FACTORS = [-7.37, 3.21, 2.10, -1.63, 1.22, -3.16, 4.01, 2.47, 2.32]

COLS_OF_INTEREST = [1, 4, 7, 9, 10, 23, 24, 25, 27, 31]
NORMALIZE_STARTING = 3
CORREL_FACTORS = [-7.369519723, 3.207527431, 2.098710386, -1.634276411, 1.215039434, -3.164313497, 4.014363826, 0.742103775, 2.474446861]


totals = []

for j in range(9):
  WS = WB[WB.sheetnames[j]]
  print(WB.sheetnames[j])

  data = []
  for row in range(2, 69):
    col_list = []
    for col in COLS_OF_INTEREST:
      col_list.append(WS.cell(row, col).value)
    data.append(col_list)

  min_max = []
  for col in range(NORMALIZE_STARTING, len(data[0])):
    min = data[0][col]
    max = data[0][col]
    for row in range(1, len(data)):
      val = data[row][col]
      if val < min: 
        min = val
      elif val > max:
        max = val
    min_max.append((min, max))


  for col in range(NORMALIZE_STARTING, len(data[0])):
    for row in range(len(data)):
      normalized = (data[row][col] - min_max[col-3][0]) / (min_max[col-3][1] - min_max[col-3][0])
      data[row][col] = normalized

  sort_list = []
  for row in range(len(data)):
    result = 0
    for col in range(1, len(data[row])):
      val = data[row][col]
      result += CORREL_FACTORS[col-1] * val
    sort_list.append((data[row][0], result))

  def comp(e):
    return e[1]

  sort_list.sort(reverse=True, key=comp)

  final_list = []
  for y in range(10):
    final_list.append((sort_list[y][0]).strip())

  print(final_list)
  name_col = 1
  cind_col = 4
  wins_col = 6

  total = 0
  for x in range(10):
    team_rank = 10 - x
    for row_num in range(1, 69):
      if WS.cell(row=row_num, column=name_col).value.strip() == final_list[x]:
        wins = WS.cell(row_num, wins_col).value
        cind = WS.cell(row_num, cind_col).value
        total += team_rank * wins
        if cind == 1:
          total += (5 + wins * 5) * wins / 2
  totals.append(total)
  print(total)

print(st.mean(totals))
print(st.stdev(totals))