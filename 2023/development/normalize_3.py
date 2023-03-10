from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv
import os

load_dotenv()
WORKBOOK_NAME = os.environ.get("WORKBOOK_NAME")
data = load_workbook(WORKBOOK_NAME)

compare_col = 6
normal_cols = [4, 7]
normalize_cols = [9, 10, 21, 23, 24, 25, 27, 31, 32]
row_range = [2, 69]
starting_col = 34

def copy_title(source_col, target_col, ws):
  ws.cell(1, target_col).value = ws.cell(1, source_col).value


for i in range(9):
  col_total = starting_col-1
  ws = data[data.sheetnames[i]]
  for col_n in normal_cols:
    col_total += 1
    copy_title(col_n, col_total, ws)
    for row_n in range(row_range[0], row_range[1]+1):
      ws.cell(row_n, col_total).value = ws.cell(row_n, col_n).value

  for col in normalize_cols:
    col_total += 1
    column_letter = get_column_letter(col)
    range_str = column_letter + str(row_range[0]) + ":" + column_letter + str(row_range[1])
    copy_title(col, col_total, ws)
    for row in range(row_range[0], row_range[1]+1):
      val = column_letter + str(row)
      formula = "=(" + val + "-MIN(" + range_str + "))/(MAX(" + range_str + ")-MIN(" + range_str + "))"
      ws.cell(row, col_total).value = formula
  
  col_total += 1
  copy_title(compare_col, col_total, ws)
  for row_c in range(row_range[0], row_range[1]+1):
    ws.cell(row_c, col_total).value = ws.cell(row_c, compare_col).value


data.save(WORKBOOK_NAME)