from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv
import os

load_dotenv()
WORKBOOK_NAME = os.environ.get("WORKBOOK_NAME")
data = load_workbook(WORKBOOK_NAME)
additional_analysis = data["Additional Analysis"]

def copy_title(source_col, target_col, ws):
  ws.cell(1, target_col).value = ws.cell(1, source_col).value

x_axis = [34, 35, 36, 37, 38, 39, 40, 41, 42, 43, 44]
y_axis = 45
starting_col = 46
row_range = [2, 69]



for i in range(9):
  ws = data[data.sheetnames[i]]
  current_col = starting_col-1
  for x in x_axis:
    col_letter = get_column_letter(x)
    y_col_letter = get_column_letter(y_axis)
    current_col += 1
    copy_title(x, current_col, ws)
    ws.cell(2, current_col).value = f"=SLOPE({col_letter}{row_range[0]}:{col_letter}{row_range[1]}, {y_col_letter}{row_range[0]}:{y_col_letter}{row_range[1]})"

data.save(WORKBOOK_NAME)
