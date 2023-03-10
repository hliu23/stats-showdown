from openpyxl import load_workbook
from dotenv import load_dotenv
import os

load_dotenv()
WORKBOOK_NAME = os.environ.get("WORKBOOK_NAME")

data = load_workbook(WORKBOOK_NAME, data_only=True)

analysis = data["Analysis"]

previous_cols = [4, 5, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27, 28, 29, 30, 31, 32]
previous_row_num = 70


for i in range(9):
  ws = data[data.sheetnames[i]]
  analysis.cell(i+2, 1).value = data.sheetnames[i]
  for col in previous_cols:
    analysis.cell(i+2, col).value = ws.cell(previous_row_num, col).value

data.save(WORKBOOK_NAME)