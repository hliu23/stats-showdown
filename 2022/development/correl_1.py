from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

data = load_workbook("testing.xlsx")

col_compare = 6
col_of_interest = [4, 5, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27, 28, 29, 30, 31, 32]
row_num = [2, 69]

def range_str(min_col, max_col, min_row=row_num[0], max_row=row_num[1]):
  res = get_column_letter(min_col) + str(min_row) + ":" + get_column_letter(max_col) + str(max_row)
  return res

range_compare = range_str(col_compare, col_compare)

for i in range(8):
  ws = data[data.sheetnames[i]]
  for col in col_of_interest:

    formula = "=CORREL(" + range_str(col, col) + ", " + range_compare + ")"
    ws.cell(70, col).value = formula

data.save("testing.xlsx")
