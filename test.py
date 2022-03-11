from openpyxl import load_workbook
WB = load_workbook("NCAABBStats2013-2021BDAB2021AlgorithmDevelopment.xlsx", read_only=True)

average = 0

for j in range(8):
  WS = WB[WB.sheetnames[j]]
  print(WB.sheetnames[j])

  COLS_OF_INTEREST = [1, 5, 7, 9, 10, 21, 23, 24, 25, 31, 32]
  data = []
  for row in range(2, 69):
    col_list = []
    for col in COLS_OF_INTEREST:
      col_list.append(WS.cell(row, col).value)
    data.append(col_list)

  min_max = []
  for col in range(3, len(data[0])):
    min = data[0][col]
    max = data[0][col]
    for row in range(1, len(data)):
      val = data[row][col]
      if val < min: 
        min = val
      elif val > max:
        max = val
    min_max.append((min, max))


  for col in range(3, len(data[0])):
    for row in range(len(data)):
      normalized = (data[row][col] - min_max[col-3][0]) / (min_max[col-3][1] - min_max[col-3][0])
      data[row][col] = normalized

  CORREL_FACTORS = [-0.17, 0.3, 0.37, -0.33, 0.22, 0.28, -0.38, 0.45, 0.44, 0.43]

  sort_list = []
  for row in range(len(data)):
    result = 0
    for col in range(1, len(data[row])):
      val = data[row][col]
      result += CORREL_FACTORS[col-1] * val
    sort_list.append((data[row][0], result))

  def comp(e):
    return e[1]

  sort_list.sort(reverse=True, key=comp)

  final_list = []
  for y in range(10):
    final_list.append((sort_list[y][0]).strip())

  print(final_list)
  name_col = 1
  cind_col = 4
  wins_col = 6

  total = 0
  

  for x in range(10):
    team_rank = 10 - x
    for row_num in range(1, WS.max_row):
      if WS.cell(row=row_num, column=name_col).value.strip() == final_list[x]:
        wins = WS.cell(row_num, wins_col).value
        cind = WS.cell(row_num, cind_col).value
        total += team_rank * wins
        if cind == 1:
          total += (5 + wins * 5) * wins / 2
  average += total
  print(total)

average /= 8
print(average)


