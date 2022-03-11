# TO RUN: install python and openpyxl package
from openpyxl import load_workbook
import sys

# TO RUN: replace workbook name and worksheet name
WB = load_workbook("<PLACEHOLDER_WORKBOOK_NAME>.xlsx", read_only=True)
WS = WB["<PLACEHOLDER_WORKSHEET_NAME"]

# TO RUN: execute with "py algorithm_downloadable.py"

# load data from workbook into nested list
COLS_OF_INTEREST = [1, 5, 7, 9, 10, 21, 23, 24, 25, 31, 32]
data = []
for row in range(2, 69):
  col_list = []
  for col in COLS_OF_INTEREST:
    col_list.append(WS.cell(row, col).value)
  data.append(col_list)

# find max and min of a column
min_max = []
for col in range(3, len(data[0])):
  min = data[0][col]
  max = data[0][col]
  for row in range(1, len(data)):
    val = data[row][col]
    if val < min: 
      min = val
    elif val > max:
      max = val
  min_max.append((min, max))


# normalize data to number between 0 and 1 if applicable
for col in range(3, len(data[0])):
  for row in range(len(data)):
    normalized = (data[row][col] - min_max[col-3][0]) / (min_max[col-3][1] - min_max[col-3][0])
    data[row][col] = normalized

# multiply data by correlation efficients calculated ahead of time and sum results for a row
CORREL_FACTORS = [-0.17, 0.3, 0.37, -0.33, 0.22, 0.28, -0.38, 0.45, 0.44, 0.43]

sort_list = []
for row in range(len(data)):
  result = 0
  for col in range(1, len(data[row])):
    val = data[row][col]
    result += CORREL_FACTORS[col-1] * val
  sort_list.append((data[row][0], result))

# sort the resultant list
def comp(e):
  return e[1]

sort_list.sort(reverse=True, key=comp)
for i in range(10):
  print(sort_list[i][0])

